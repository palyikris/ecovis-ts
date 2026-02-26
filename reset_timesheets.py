# -*- coding: utf-8 -*-
"""
TS reset tool
- Minden TS *.xlsx fájlt archivál, majd ugyanazzal a névvel "üres" példányt hoz létre,
  megőrizve a formátumot és az érvényesítéseket.
- A hónap-lapokon A2..X301 tartományt ürít (Y/Z segédoszlopok, validációk megmaradnak).

Usage:
    python reset_timesheets.py
    python reset_timesheets.py --dry-run
    python reset_timesheets.py --folder . --max-rows 300 --clear-until-col X
"""
from __future__ import annotations

import argparse
import logging
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# --- Konfiguráció (alapértékek) ---
DEFAULT_FOLDER = "."
DEFAULT_MAX_ROWS = 300  # projekt konvenció (scripts is 300 sorral dolgoznak)
DEFAULT_CLEAR_UNTIL_COL = "X"  # Y/Z segédoszlopok meghagyása (drop-down források)

# Magyar hónapok (ékezet nélkül) – a projekttel konzisztensen
HONAPOK = [
    "januar",
    "februar",
    "marcius",
    "aprilis",
    "majus",
    "junius",
    "julius",
    "augusztus",
    "szeptember",
    "oktober",
    "november",
    "december",
]


def remove_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s or "")
    return "".join(c for c in nfkd if not unicodedata.combining(c)).strip().lower()


def is_ts_file(name: str) -> bool:
    # Konzisztens szűrés: .xlsx és tartalmazza a "TS" mintát, nem ideiglenes (~$)
    return name.endswith(".xlsx") and "TS" in name and not name.startswith("~$")


def make_archive_dir(base: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    d = base / f"archived_ts_{ts}"
    d.mkdir(parents=True, exist_ok=True)
    return d


def clear_month_sheet(ws, max_rows: int, clear_until_col_letter: str) -> None:
    """Üríti a lap A2..{clear_until_col_letter}{max_rows} tartományát.
    A Y/Z segédoszlopok érintetlenek maradnak, így a drop-down forráslisták megmaradnak.
    """
    max_col_idx = column_index_from_string(clear_until_col_letter)
    # fejléc: 1. sor, adatok: 2..max_rows
    for r in ws.iter_rows(min_row=2, max_row=max_rows, min_col=1, max_col=max_col_idx):
        for cell in r:
            cell.value = None


def create_blank_from_archived(
    archived_path: Path, new_path: Path, max_rows: int, clear_until_col_letter: str
) -> None:
    # Archív példányból olvasunk, hogy az új fájl szerkezete/validációi 1:1-ben megmaradjanak
    wb = load_workbook(archived_path, data_only=False)
    for ws in wb.worksheets:
        name_norm = remove_accents(ws.title)
        if name_norm in HONAPOK:
            clear_month_sheet(ws, max_rows, clear_until_col_letter)
            # Megjegyzés: a data_validations, freeze_panes, oszlopszélességek a munkalap szintjén
            # maradnak, mi csak a cellaértékeket töröljük.
    wb.save(new_path)


def main():
    parser = argparse.ArgumentParser(description="Reset TS workbooks safely.")
    parser.add_argument(
        "--folder",
        "-f",
        default=DEFAULT_FOLDER,
        help="Mappa, ahol a TS fájlok vannak (default: .)",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=DEFAULT_MAX_ROWS,
        help="Törlendő sorok száma (alap: 300)",
    )
    parser.add_argument(
        "--clear-until-col",
        default=DEFAULT_CLEAR_UNTIL_COL,
        help="Utolsó törlendő oszlop betűjele (alap: X)",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Csak listáz, nem módosít"
    )
    args = parser.parse_args()

    folder = Path(args.folder).resolve()
    max_rows = args.max_rows
    clear_until_col = args.clear_until_col.upper()

    log_dir = folder / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = (
        log_dir / f"reset_timesheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    logging.info("▶ reset_timesheets started")
    logging.info(f"Folder: {folder}")
    logging.info(f"Log file: {log_file}")

    # Cél archív mappa
    archive_dir = make_archive_dir(folder)
    logging.info(f"Archive dir: {archive_dir}")

    ts_files = [p for p in folder.iterdir() if p.is_file() and is_ts_file(p.name)]
    if not ts_files:
        logging.info("Nincs feldolgozható TS fájl.")
        return

    moved = 0
    created = 0
    errors = 0

    for src in ts_files:
        try:
            dst_arch = archive_dir / src.name
            logging.info(f"🔧 Feldolgozás: {src.name}")

            if args.dry_run:
                logging.info(f"   ↪ DRY-RUN: move -> {dst_arch.name}")
                logging.info(f"   ↪ DRY-RUN: recreate blank -> {src.name}")
                moved += 1
                created += 1
                continue

            # 1) Átmozgatás archívba
            shutil.move(str(src), str(dst_arch))
            moved += 1
            logging.info(f"   ✔ Áthelyezve: {dst_arch}")

            # 2) Üres példány létrehozása (azonos szerkezet/validációk)
            create_blank_from_archived(dst_arch, src, max_rows, clear_until_col)
            created += 1
            logging.info(f"   ✔ Új üres fájl létrehozva: {src.name}")

        except Exception as e:
            errors += 1
            logging.exception(f"❌ Hiba: {src.name} — {e}")

    logging.info("📊 Összegzés:")
    logging.info(f"   ➜ Áthelyezett fájlok: {moved}")
    logging.info(f"   ➜ Létrehozott üres fájlok: {created}")
    logging.info(f"   ➜ Hibák: {errors}")
    logging.info("✅ reset_timesheets finished")


if __name__ == "__main__":
    main()
