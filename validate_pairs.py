# validate_pairs.py
import pandas as pd
import os
import sys
import re
import unicodedata
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

import logging
from pathlib import Path
import time

# --- LOGGING ---
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"validate_pairs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()],
)
logging.info("validate_pairs started")
logging.info(f"Log file: {LOG_FILE.resolve()}")

# --- CONFIG ---
FOLDER_PATH = "."
ECOVIS_PATH = "Ecovis Compliance Solution számlázási adatok_2025.xlsx"
TS_KODOK_SHEET = "TS kódok"
MAX_ROWS = 300
SKIP_MISSING = (
    False  # ha False, a hiányzó Ügyfélkód/Projekt neve sorok is bekerülnek a riportba
)
LOGO_PATH = "ecovis_logo.png"  # ha ott van a mappában, betesszük a fejlécbe
BRAND_COLOR = "D92D27"  # corporate piros (fejléc sáv)
ACCENT_COLOR = "4F81BD"  # kék akcentus (összegzés fejlécekhez stb.)

# --- Counters ---
start_time = time.time()
processed_files = 0
skipped_files = 0
errors = 0
row_issues_total = 0


# --- Helpers ---
def remove_accents(s: str) -> str:
    if not isinstance(s, str):
        s = "" if pd.isna(s) else str(s)
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).strip().lower()


HONAPOK = [
    "januar",
    "februar",
    "marcius",
    "aprilis",
    "majus",
    "junius",
    "julius",
    "augusztus",
    "szeptember",
    "oktober",
    "november",
    "december",
]


def resolve_selected_month(arg_month: Optional[str]) -> str:
    if arg_month:
        m = remove_accents(arg_month)
        logging.info(f"Hónap paraméterből: {m}")
        return m
    now = datetime.now()
    m = HONAPOK[now.month - 1]
    logging.info(f"Hónap alapértelmezett (aktuális): {m}")
    return m


def load_allowed_map() -> dict[str, set[str]]:
    logging.info(f"TS kódok beolvasása: {ECOVIS_PATH} / {TS_KODOK_SHEET}")
    df = pd.read_excel(
        ECOVIS_PATH, sheet_name=TS_KODOK_SHEET, usecols=["Ügyfélkód", "Projekt neve"]
    )
    df = df.dropna(subset=["Ügyfélkód", "Projekt neve"])
    df["kod_norm"] = df["Ügyfélkód"].astype(str).map(remove_accents)
    df["prj_norm"] = df["Projekt neve"].astype(str).map(remove_accents)
    allowed: dict[str, set[str]] = {}
    for _, r in df.iterrows():
        allowed.setdefault(r["kod_norm"], set()).add(r["prj_norm"])
    logging.info(
        f"Engedélyezett párok betöltve: {len(allowed)} ügyfélkód, összesen ~{int(df.shape[0])} sor"
    )
    return allowed


def validate_file(
    ts_path: str, month_norm: str, allowed: dict[str, set[str]]
) -> list[list]:
    rows: list[list] = []
    basename = os.path.basename(ts_path)
    logging.info(f"Feldolgozás: {basename}")
    try:
        xls = pd.ExcelFile(ts_path)
    except Exception as e:
        rows.append([basename, "-", "-", "-", "-", f"Nem nyitható: {e}"])
        logging.exception(f"Nem nyitható: {basename} — {e}")
        return rows

    # keresett hónap sheet
    target_sheet = None
    for s in xls.sheet_names:
        if remove_accents(s) == month_norm:
            target_sheet = s
            break
    if not target_sheet:
        logging.info(f"Kihagyva (nincs megfelelő hónap sheet): {basename}")
        return rows  # nincs ilyen sheet -> nincs mit ellenőrizni

    logging.info(f"  Sheet: {target_sheet}")

    usecols = ["Ügyfélkód", "Projekt neve"]
    try:
        df = pd.read_excel(
            xls, sheet_name=target_sheet, usecols=usecols, nrows=MAX_ROWS, dtype=str
        )
    except Exception as e:
        rows.append(
            [basename, target_sheet, "-", "-", "-", f"Sheet olvasási hiba: {e}"]
        )
        logging.exception(f"Sheet olvasási hiba ({basename}/{target_sheet}): {e}")
        return rows

    for idx, r in df.iterrows():
        kod_raw = r.get("Ügyfélkód", None)
        prj_raw = r.get("Projekt neve", None)
        excel_row = idx + 2  # A1 fejlec, adatok 2-től
        # load active clients once globally (top of file)
        ceg = pd.read_excel(ECOVIS_PATH, sheet_name="Cégadatok")
        active_clients = set(
            ceg[ceg["Ügyfél aktív"].astype(str).str.strip().str.lower() == "igen"]["Ügyfélkód"].astype(str).map(remove_accents)
        )
        
        if kod_norm not in active_clients:
            continue  # passzív ügyfél – TS sor kihagyva teljesen


        kod_norm = remove_accents(kod_raw)
        prj_norm = remove_accents(prj_raw)
        
        if kod_norm == "eco":
            continue  # ECO kódot nem ellenőrzünk

        if SKIP_MISSING and (kod_norm == "" or prj_norm == ""):
            continue

        if kod_norm == "" and prj_norm == "":
            # rows.append(
            #     [
            #         basename,
            #         target_sheet,
            #         excel_row,
            #         str(kod_raw or ""),
            #         str(prj_raw or ""),
            #         "Hiányzó Ügyfélkód és Projekt neve",
            #     ]
            # )
            continue
        if kod_norm == "":
            rows.append(
                [
                    basename,
                    target_sheet,
                    excel_row,
                    "#" if pd.isna(kod_raw) or kod_raw == "" else str(kod_raw),
                    "#" if pd.isna(prj_raw) or prj_raw == "" else str(prj_raw),
                    "Hiányzó Ügyfélkód",
                ]
            )
            continue
        if prj_norm == "":
            rows.append(
                [
                    basename,
                    target_sheet,
                    excel_row,
                    "#" if pd.isna(kod_raw) or kod_raw == "" else str(kod_raw),
                    "#" if pd.isna(prj_raw) or prj_raw == "" else str(prj_raw),
                    "Hiányzó Projekt neve",
                ]
            )
            continue

        if kod_norm not in allowed:
            rows.append(
                [
                    basename,
                    target_sheet,
                    excel_row,
                    str(kod_raw),
                    str(prj_raw),
                    "Ismeretlen Ügyfélkód (nincs a TS kódokban)",
                ]
            )
            continue

        if prj_norm not in allowed[kod_norm]:
            rows.append(
                [
                    basename,
                    target_sheet,
                    excel_row,
                    str(kod_raw),
                    str(prj_raw),
                    "Érvénytelen páros: Ügyfélkódhoz ez a Projekt nem engedélyezett",
                ]
            )

    if rows:
        logging.info(f"Hibás sorok a fájlban: {len(rows)}")
    else:
        logging.info("Nincs hiba ebben a fájlban")
    return rows


# --- Excel styling helpers ---
def autosize_columns(
    ws,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
):
    """MergedCell-safe: indexekből kalkulálja az oszlopbetűt, nem cellából.
    min_row-t érdemes a táblázat fejlécre állítani (pl. 4), hogy a banner ne torzítson.
    """
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column
    for col_idx in range(min_col, max_col + 1):
        max_len = 0
        for row in range(min_row, max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is not None:
                l = len(str(v))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def add_title_banner(ws, title: str, subtitle: str):
    # címsor a1..f1, piros sáv
    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=BRAND_COLOR)

    ws.merge_cells("A2:F2")
    ws["A2"].value = subtitle
    ws["A2"].font = Font(size=11, color="333333")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def write_table(
    ws,
    start_row: int,
    df: pd.DataFrame,
    table_name: str,
    table_style: str = "TableStyleMedium9",
):
    # fejlécek
    headers = list(df.columns)
    ws.append([""] * len(headers))  # spacer
    row0 = start_row
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row0, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor=ACCENT_COLOR)

    # adatok
    for _, r in df.iterrows():
        ws.append([r.get(h) for h in headers])

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(headers))
    ref = f"A{row0}:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name=table_style, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    # fagyasztás a fejléc alatt
    ws.freeze_panes = None


def add_error_highlights(ws, header_row: int):
    # "Hiba" oszlop megkeresése
    col_idx = None
    for cell in ws[header_row]:
        if (str(cell.value or "")).strip().lower() == "hiba":
            col_idx = cell.column
            break
    if col_idx is None:
        return

    hcol = get_column_letter(col_idx)
    rng = f"{hcol}{header_row+1}:{hcol}{ws.max_row}"

    # piros háttér "Érvénytelen páros" esetén
    red = DifferentialStyle(fill=PatternFill("solid", fgColor="F8D7DA"))
    rule1 = Rule(
        type="containsText",
        operator="containsText",
        text="Érvénytelen páros",
        dxf=red,
        stopIfTrue=False,
    )
    rule1.formula = [f'NOT(ISERROR(SEARCH("Érvénytelen páros",{hcol}{header_row+1})))']
    ws.conditional_formatting.add(rng, rule1)

    # sárga háttér "Hiányzó" esetén
    yellow = DifferentialStyle(fill=PatternFill("solid", fgColor="FFF3CD"))
    rule2 = Rule(
        type="containsText",
        operator="containsText",
        text="Hiányzó",
        dxf=yellow,
        stopIfTrue=False,
    )
    rule2.formula = [f'NOT(ISERROR(SEARCH("Hiányzó",{hcol}{header_row+1})))']
    ws.conditional_formatting.add(rng, rule2)


def build_summary_sheet(wb: Workbook, data_df: pd.DataFrame, month_txt: str):
    ws = wb.create_sheet("Összegzés")
    add_title_banner(
        ws,
        "Ügyfélkód–Projekt párok ellenőrzése — Összegzés",
        f"Hónap: {month_txt}    Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )

    start = 4
    # 1) Összes hiba darabszám
    ws["A" + str(start)].value = "Összes hibás sor"
    ws["A" + str(start)].font = Font(bold=True)
    ws["B" + str(start)].value = len(data_df.index)
    start += 2

    # 2) Hibatípusok számossága
    if not data_df.empty:
        by_err = (
            data_df.groupby("Hiba")
            .size()
            .reset_index(name="Darab")
            .sort_values("Darab", ascending=False)
        )
        for _ in range(1):  # spacer
            ws.append([])
        ws["A" + str(start)].value = "Hibatípusok"
        ws["A" + str(start)].font = Font(bold=True)
        write_table(
            ws, start + 1, by_err, "Hibatipusok", table_style="TableStyleMedium2"
        )
        start = ws.max_row + 2

        # 3) Hibák fájlonként
        by_file = (
            data_df.groupby("Fájl")
            .size()
            .reset_index(name="Darab")
            .sort_values("Darab", ascending=False)
        )
        ws["A" + str(start)].value = "Hibák fájlonként"
        ws["A" + str(start)].font = Font(bold=True)
        write_table(
            ws, start + 1, by_file, "HibakFajlonkent", table_style="TableStyleMedium4"
        )
        start = ws.max_row + 2

        # 4) Ismétlődő hibás párok
        if {"Ügyfélkód", "Projekt neve"}.issubset(set(data_df.columns)):
            by_pair = (
                data_df.groupby(["Ügyfélkód", "Projekt neve"])
                .size()
                .reset_index(name="Darab")
                .sort_values("Darab", ascending=False)
            )
            ws["A" + str(start)].value = "Ismétlődő hibás párok"
            ws["A" + str(start)].font = Font(bold=True)
            write_table(
                ws, start + 1, by_pair, "HibasParok", table_style="TableStyleMedium9"
            )

    # Ne legyen fagyasztás az Összegzés lapon:
    ws.freeze_panes = None

    # A banner miatt jobb, ha a táblázatoktól méretezünk
    autosize_columns(ws, min_row=4)


def main():
    global processed_files, skipped_files, errors, row_issues_total

    try:
        selected_month = resolve_selected_month(
            sys.argv[1] if len(sys.argv) > 1 else None
        )
        month_txt = selected_month  # ékezetmentes név

        # Engedélyezett párosok
        allowed = load_allowed_map()

        # Ellenőrzés
        all_rows: list[list] = []
        for fname in os.listdir(FOLDER_PATH):
            if fname.endswith(".xlsx") and "TS" in fname and not fname.startswith("~$"):
                print(f"🔧 Feldolgozás: {fname}")
                path = os.path.join(FOLDER_PATH, fname)
                before = len(all_rows)
                rows = validate_file(path, selected_month, allowed)
                all_rows.extend(rows)
                if rows is None:
                    skipped_files += 1
                else:
                    processed_files += 1

        row_issues_total = len(all_rows)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_name = f"invalid_parok_{month_txt}_{ts}.xlsx"

        # Workbook & szépítés
        wb = Workbook()
        ws = wb.active
        ws.title = "Hibák"

        add_title_banner(
            ws,
            "Ügyfélkód–Projekt párok ellenőrzése — Hibák",
            f"Hónap: {month_txt}    Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        # Dataframe a hibákról (vagy üres táblázat fallback)
        if not all_rows:
            df = pd.DataFrame(
                columns=["Fájl", "Hónap", "Sor", "Ügyfélkód", "Projekt neve", "Hiba"]
            )
        else:
            df = pd.DataFrame(
                all_rows,
                columns=["Fájl", "Hónap", "Sor", "Ügyfélkód", "Projekt neve", "Hiba"],
            )
            df.sort_values(by=["Fájl", "Hónap", "Sor"], inplace=True, kind="stable")

        # táblázat beírása és formázása
        start_row = 4  # fejléc sáv után
        write_table(ws, start_row, df, "HibasSorok", table_style="TableStyleMedium9")
        add_error_highlights(ws, header_row=start_row)
        autosize_columns(ws, min_row=start_row)

        # Összegző sheet
        build_summary_sheet(wb, df, month_txt)

        wb.save(out_name)
        if df.empty:
            msg = f"Nincs hiba. Üres, de formázott jelentés készült: {out_name}"
            print(msg)
            logging.info(msg)
        else:
            msg = f"Kész a formázott hibalista: {out_name}"
            print(msg)
            logging.info(msg)

    except Exception as e:
        errors += 1
        logging.exception(f"Végzetes hiba futás közben: {e}")
        raise
    finally:
        duration = time.time() - start_time
        logging.info("Run summary:")
        logging.info(f"   {processed_files} files processed")
        logging.info(f"   {skipped_files} files skipped")
        logging.info(f"   {errors} errors")
        logging.info(f"   Hibás sorok összesen: {row_issues_total}")
        logging.info(f"   Duration: {duration:.1f}s")
        logging.info("validate_pairs finished")


if __name__ == "__main__":
    main()
