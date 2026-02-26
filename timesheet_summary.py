# -*- coding: utf-8 -*-
# timesheet_summary.py — Aggregált, “céges” kimenet (fagyasztás és logó nélkül)
import pandas as pd
import os
import re
import unicodedata
import sys
from datetime import datetime
import logging
from pathlib import Path
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# -------------------------
# Config
# -------------------------
FOLDER_PATH = "."
MAX_ROWS_PER_SHEET = 300
BRAND_COLOR = "D92D27"  # fejléc sáv
ACCENT_COLOR = "4F81BD"  # táblázat fejléc

# -------------------------
# Logging (UTF-8)
# -------------------------
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"timesheet_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logging.info("▶ timesheet_summary started")
logging.info(f"Log file: {LOG_FILE.resolve()}")


# -------------------------
# Helpers
# -------------------------
def remove_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def norm_header(s: str) -> str:
    """Oszlopnév normalizálás: ékezet nélkül, kisbetű, környezeti szóközök nélkül."""
    return remove_accents((s or "")).strip().lower()


HONAPOK = [
    "januar",
    "februar",
    "marcius",
    "aprilis",
    "majus",
    "junius",
    "julius",
    "augusztus",
    "szeptember",
    "oktober",
    "november",
    "december",
]
honap_regex = re.compile("^(" + "|".join(HONAPOK) + ")$", re.IGNORECASE)


def autosize_columns(ws, min_row: int = 1, min_col: int = 1):
    max_row = ws.max_row
    max_col = ws.max_column
    for col_idx in range(min_col, max_col + 1):
        max_len = 0
        for r in range(min_row, max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def add_title_banner(ws, title: str, subtitle: str):
    # sáv (A1..F1) — logó NINCS, fagyasztás NINCS
    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=BRAND_COLOR)

    ws.merge_cells("A2:F2")
    ws["A2"].value = subtitle
    ws["A2"].font = Font(size=11, color="333333")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def write_table(
    ws,
    start_row: int,
    df: pd.DataFrame,
    table_name: str,
    table_style: str = "TableStyleMedium9",
):
    """Filterezhető Excel-táblázat létrehozása fagyasztás nélkül."""
    headers = list(df.columns)
    if start_row > 1:
        ws.append([""] * len(headers))
    row0 = start_row

    # fejléc sor
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row0, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor=ACCENT_COLOR)

    # adatsorok
    for _, r in df.iterrows():
        ws.append([r.get(h) for h in headers])

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(headers))
    ref = f"A{row0}:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name=table_style, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    # kifejezett kérés: NE legyen fagyasztás
    ws.freeze_panes = None


def add_hour_highlights(ws, header_row: int, col_name: str = "Óra"):
    """Finom kiemelés az órákra (0 → sárga, 160 felett → halvány piros)."""
    # keresd a 'Óra' oszlopot
    col_idx = None
    for cell in ws[header_row]:
        if (str(cell.value or "")).strip().lower() == col_name.lower():
            col_idx = cell.column
            break
    if not col_idx:
        return
    hcol = get_column_letter(col_idx)
    rng = f"{hcol}{header_row+1}:{hcol}{ws.max_row}"

    yellow = DifferentialStyle(fill=PatternFill("solid", fgColor="FFF3CD"))
    rule_zero = Rule(type="cellIs", operator="equal", dxf=yellow, formula=["0"])
    ws.conditional_formatting.add(rng, rule_zero)

    light_red = DifferentialStyle(fill=PatternFill("solid", fgColor="F8D7DA"))
    rule_high = Rule(
        type="cellIs", operator="greaterThan", dxf=light_red, formula=["160"]
    )
    ws.conditional_formatting.add(rng, rule_high)


def find_description_column(columns: list[str]) -> str | None:
    """
    Megkeresi a 'munka leírása' oszlopot több alias alapján.
    Visszaadja az eredeti oszlopnevet (nem normalizált), ha talál.
    """
    aliases = [
        "munka leirasa",
        "munka leírása",
        "leiras",
        "leírás",
        "megjegyzes",
        "megjegyzés",
        "feladat leirasa",
        "feladat leírása",
        "feladat",
        "tevékenység",
        "tevekenyseg",
        "munka",
    ]
    norm_map = {norm_header(c): c for c in columns}
    for a in aliases:
        if a in norm_map:
            return norm_map[a]
    return None


# -------------------------
# Arg: hónap
# -------------------------
if len(sys.argv) > 1:
    selected_month_raw = sys.argv[1]
else:
    selected_month_raw = None

if selected_month_raw and selected_month_raw.lower() != "teljes év":
    month_norm = remove_accents(selected_month_raw.lower())
    month_label = month_norm
    logging.info(f"Hónap szűrő: {month_norm}")
else:
    month_norm = None
    month_label = "teljes_ev"
    logging.info("Hónap szűrő: TELJES ÉV")


# --- load active clients from Cégadatok ---
ceg = pd.read_excel(
    "Ecovis Compliance Solution számlázási adatok_2025.xlsx", sheet_name="Cégadatok"
)
active_clients = set(
    ceg[ceg["Ügyfél aktív"].astype(str).str.strip().str.lower() == "igen"][
        "Ügyfélkód"
    ].astype(str)
)


# -------------------------
# Gyűjtés
# -------------------------
records: list[dict] = []

start_time = time.time()
processed_files = 0
skipped_files = 0
errors = 0
processed_sheets = 0
skipped_sheets = 0

for file in os.listdir(FOLDER_PATH):
    if file.endswith(".xlsx") and "TS" in file and not file.startswith("~$"):
        file_path = os.path.join(FOLDER_PATH, file)
        logging.info(f"🔧 Feldolgozás: {file}")
        try:
            xls = pd.ExcelFile(file_path)
        except Exception as e:
            errors += 1
            logging.exception(f"❌ Nem sikerült megnyitni: {file} — {e}")
            continue

        had = False
        for sheet in xls.sheet_names:
            s_norm = norm_header(sheet)
            if month_norm:
                if s_norm != month_norm:
                    skipped_sheets += 1
                    continue
            else:
                if not honap_regex.fullmatch(s_norm):
                    skipped_sheets += 1
                    continue

            had = True
            logging.info(f"  ➔ Sheet: {sheet}")

            # Olvassuk be a lapot teljes oszlopkészlettel (robosztusabb a leírás oszlop variánsaira)
            try:
                df = pd.read_excel(
                    xls,
                    sheet_name=sheet,
                    nrows=MAX_ROWS_PER_SHEET,
                )
            except Exception as e:
                errors += 1
                logging.exception(
                    f"    ❌ Hiba a sheet olvasásakor ({file}/{sheet}): {e}"
                )
                continue

            df.dropna(how="all", inplace=True)
            if df.empty:
                logging.info("    ➔ Üres sheet, kihagyva")
                skipped_sheets += 1
                continue

            # Szükséges “kötelező” oszlopok
            needed = ["Ügyfélkód", "Projekt neve", "Időráfordítás (óra)"]
            # Tudd meg a (változó nevű) leírás oszlopot
            desc_col = find_description_column(list(df.columns))

            # Ellenőrizd a kötelező oszlopokat
            miss = [c for c in needed if c not in df.columns]
            if miss:
                logging.warning(f"    ➔ Hiányzó oszlop(ok): {miss}, kihagyva")
                skipped_sheets += 1
                continue

            # dolgozó (fájlnév)
            person = file.replace(".xlsx", "")

            # Csak komplett sorok (óra, ügyfélkód, projekt név)
            df = df.dropna(subset=["Ügyfélkód", "Projekt neve", "Időráfordítás (óra)"])
            if df.empty:
                skipped_sheets += 1
                continue

            for _, r in df.iterrows():
                try:
                    hours = float(r["Időráfordítás (óra)"])
                except Exception:
                    continue
                # Leírás érték (ha nincs oszlop, akkor üres string)
                desc_val = ""
                if desc_col is not None:
                    val = r.get(desc_col, "")
                    desc_val = "" if pd.isna(val) else str(val)
                    
                kod = str(r["Ügyfélkód"])
                
                if kod not in active_clients:
                    continue


                records.append(
                    {
                        "Ügyfélkód": str(r["Ügyfélkód"]),
                        "Projekt neve": str(r["Projekt neve"]),
                        "Munka leírása": desc_val,
                        "Dolgozó": person,
                        "Forrás fájl": file,  # <--- ÚJ: konkrét TS fájlnév
                        "Óra": round(hours, 2),
                    }
                )
            processed_sheets += 1

        if had:
            processed_files += 1
        else:
            skipped_files += 1
            logging.info(f"⚠️ Kihagyva (nincs releváns hónap sheet): {file}")

# -------------------------
# DataFrames
# -------------------------
if records:
    df_long = pd.DataFrame.from_records(
        records,
        columns=["Ügyfélkód", "Projekt neve", "Munka leírása", "Dolgozó", "Forrás fájl", "Óra"],
    )
else:
    df_long = pd.DataFrame(
        columns=["Ügyfélkód", "Projekt neve", "Munka leírása", "Dolgozó", "Forrás fájl", "Óra"]
    )

# 1) AGGREGÁLT első lap
#    (Ügyfélkód + Projekt neve + Munka leírása → össz. óra, és a forrás fájlok listája)
if df_long.empty:
    df_agg = pd.DataFrame(columns=["Ügyfélkód", "Projekt neve", "Munka leírása", "Óra", "Forrás fájl(ok)"])
else:
    grouped = (
        df_long.groupby(["Ügyfélkód", "Projekt neve", "Munka leírása"], dropna=False)
        .agg(
            Óra=("Óra", "sum"),
            _forras=("Forrás fájl", lambda s: ", ".join(sorted(set(map(str, s)))))
        )
        .reset_index()
        .sort_values(["Ügyfélkód", "Projekt neve", "Munka leírása"], kind="stable")
    )
    df_agg = grouped.rename(columns={"_forras": "Forrás fájl(ok)"})

# 2) Nézetek: Dolgozónként
by_person = (
    df_long.groupby(["Dolgozó"], dropna=False)["Óra"]
    .sum()
    .reset_index()
    .sort_values(["Óra"], ascending=False, kind="stable")
)

# Top projektek (összóra szerint) — leírástól függetlenül
top_projects = (
    df_long.groupby(["Ügyfélkód", "Projekt neve"], dropna=False)["Óra"]
    .sum()
    .reset_index()
    .sort_values("Óra", ascending=False, kind="stable")
    .head(20)
    .reset_index(drop=True)
)

# -------------------------
# Excel kiírás (fagyasztás/logó nélkül)
# -------------------------
out_name = (
    f"timesheet_summary_{month_label}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
)
wb = Workbook()

# Összesítés (aggregált) — első lap
ws_main = wb.active
ws_main.title = "Összesítés"

add_title_banner(
    ws_main,
    f"Timesheet összesítés — {month_label}",
    f"Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
)

# Ha üres, akkor is legyen fejléces tábla
if df_agg.empty:
    tbl_df = pd.DataFrame(columns=["Ügyfélkód", "Projekt neve", "Munka leírása", "Óra", "Forrás fájl(ok)"])
else:
    tbl_df = df_agg

write_table(ws_main, start_row=4, df=tbl_df, table_name="Osszesites")
add_hour_highlights(ws_main, header_row=4, col_name="Óra")
autosize_columns(ws_main, min_row=4)

# Nézetek lap (szűrhető táblázatok)
ws_views = wb.create_sheet("Nézetek")
add_title_banner(
    ws_views,
    f"Nézetek — {month_label}",
    f"Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
)

# (a) Dolgozónként
ws_views["A4"].value = "Összesítés dolgozónként"
ws_views["A4"].font = Font(bold=True)
df_person_tbl = by_person.rename(columns={"Óra": "Óra"})
write_table(
    ws_views,
    start_row=5,
    df=df_person_tbl,
    table_name="ByPerson",
    table_style="TableStyleMedium4",
)
add_hour_highlights(ws_views, header_row=5, col_name="Óra")
autosize_columns(ws_views, min_row=5)

# (b) Top projektek (leírástól függetlenül)
start2 = ws_views.max_row + 3
ws_views["A" + str(start2)].value = "Top projektek (óra szerint)"
ws_views["A" + str(start2)].font = Font(bold=True)
write_table(
    ws_views,
    start_row=start2 + 1,
    df=top_projects,
    table_name="TopProjects",
    table_style="TableStyleMedium9",
)
add_hour_highlights(ws_views, header_row=start2 + 1, col_name="Óra")
autosize_columns(ws_views, min_row=start2 + 1)

# Összegzés lap (kulcsszámok)
ws_sum = wb.create_sheet("Összegzés")
add_title_banner(
    ws_sum,
    f"Összegzés — {month_label}",
    f"Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
)
r = 4
ws_sum["A" + str(r)].value = "Feldolgozott fájlok"
ws_sum["A" + str(r)].font = Font(bold=True)
ws_sum["B" + str(r)].value = processed_files
r += 1
ws_sum["A" + str(r)].value = "Kihagyott fájlok"
ws_sum["A" + str(r)].font = Font(bold=True)
ws_sum["B" + str(r)].value = skipped_files
r += 1
ws_sum["A" + str(r)].value = "Feldolgozott sheetek"
ws_sum["A" + str(r)].font = Font(bold=True)
ws_sum["B" + str(r)].value = processed_sheets
r += 1
ws_sum["A" + str(r)].value = "Kihagyott sheetek"
ws_sum["A" + str(r)].font = Font(bold=True)
ws_sum["B" + str(r)].value = skipped_sheets
r += 2
ws_sum["A" + str(r)].value = "Összes idő (óra)"
ws_sum["A" + str(r)].font = Font(bold=True)
ws_sum["B" + str(r)].value = float(df_agg["Óra"].sum()) if not df_agg.empty else 0.0
autosize_columns(ws_sum, min_row=4)

# Mentés
try:
    wb.save(out_name)
    logging.info(f"✅ Összesítés elkészült, elmentve ide: {out_name}")
except Exception as e:
    errors += 1
    logging.exception(f"❌ Nem sikerült kiírni az eredményt: {e}")

# -------------------------
# Summary log
# -------------------------
duration = time.time() - start_time
logging.info("📊 Run summary:")
logging.info(f"   ✔ {processed_files} files processed")
logging.info(f"   ⚠ {skipped_files} files skipped (no target month)")
logging.info(f"   📄 {processed_sheets} sheets processed")
logging.info(f"   💤 {skipped_sheets} sheets skipped")
logging.info(f"   ❌ {errors} errors")
logging.info(f"   ⏱ Duration: {duration:.1f}s")
logging.info("✅ timesheet_summary finished")
