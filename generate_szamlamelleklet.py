# generate_szamlamelleklet.py
# -*- coding: utf-8 -*-
import pandas as pd
import os
import re
import unicodedata
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import sys
import logging
from pathlib import Path

# ---- CONFIG ----
FOLDER_PATH = "."
COMPLIANCE_FILE = "Ecovis Compliance Solution számlázási adatok_2025.xlsx"
CEGADATOK_SHEET = "Cégadatok"
LOGO_CANDIDATES = ["ecovis_logo.png", "/mnt/data/ecovis_logo.png"]
MAX_ROWS_PER_SHEET = 300

# ➤ Alapértelmezett (korábban kért) lista:
ORDERED_CODES_DEFAULT = [
    "AUC",
    "AXM",
    "BRD",
    "HÖG",
    "ITP",
    "JIS",
    "KKE",
    "KLU",
    "KRT",
    "LUT",
    "MES",
    "NUM",
    "OLD",
    "PCO",
    "PRM",
    "RAP",
    "ROC",
    "SCH",
    "SPA",
    "TLA",
    "VAB",
    "ZAP",
]

BOILERPLATE = {
    "magyar": {
        "title": "Számlamelléklet (teljesítési igazolás)",
        "period": "{year}. {month}. hónap",
        "date": "Budapest, {today}",
        "text": (
            "Szerződésünk 4 pontja szerint csatoljuk az adott elszámolási időszakban "
            "igénybe vett tanácsadási szolgáltatásokról szóló kimutatást."
        ),
        "task_header": "Feladat",
        "hours_header": "Időráfordítás (óra)",
        "summary_labels": [
            "Felhasznált tanácsadói órák",
            "Szerződés szerint rendelkezésre álló óraszám",
            "Korábbi időszaki órák",
            "Különbözet",
        ],
    },
    "angol": {
        "title": "Invoice attachment (certificate of completion)",
        "period": "{year}-{month}",
        "date": "Budapest, {today}",
        "text": (
            "According to point 4 of our contract,"
            "we are attaching a statement of the consulting services used in the given accounting period."
        ),
        "task_header": "Task description",
        "hours_header": "Time spent (hours)",
        "summary_labels": [
            "Used consulting hours",
            "Contracted available hours",
            "Older period hours",
            "Difference",
        ],
    },
}


# ---- LOGGING ----
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = (
    LOG_DIR / f"generate_szamlamelleklet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()],
)
logging.info("generate_szamlamelleklet started")
logging.info(f"Log file: {LOG_FILE.resolve()}")


# ---- HELPERS ----
def remove_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


HONAPOK = [
    "januar",
    "februar",
    "marcius",
    "aprilis",
    "majus",
    "junius",
    "julius",
    "augusztus",
    "szeptember",
    "oktober",
    "november",
    "december",
]
HONAP_TO_Q = {1: 1, 2: 1, 3: 1, 4: 2, 5: 2, 6: 2, 7: 3, 8: 3, 9: 3, 10: 4, 11: 4, 12: 4}

INVALID_SHEET_CHARS = r"[:\\/?*\[\]]"


def sanitize_sheet_title(raw: str, used: set[str]) -> str:
    name = re.sub(INVALID_SHEET_CHARS, "_", str(raw)).strip() or "Lap"
    if len(name) > 31:
        name = name[:31]
    base, i = name, 1
    while name in used:
        add = f"_{i}"
        name = base[: (31 - len(add))] + add
        i += 1
    used.add(name)
    return name


def find_logo_path() -> str | None:
    for p in LOGO_CANDIDATES:
        if Path(p).exists():
            return p
    return None


# --- stílusok egy oszlopos layout-hoz ---
THIN = Side(style="thin", color="999999")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
THICK_BLACK = Side(style="medium", color="000000")
BORDER_BOTTOM_THICK = Border(bottom=THICK_BLACK)


def place_logo_top_left(ws):
    lp = find_logo_path()
    if not lp:
        logging.warning("Logó nem található (ecovis_logo.png).")
        return
    try:
        img = XLImage(lp)
        target_h = 60
        if img.height > target_h:
            scale = target_h / img.height
            img.height = int(img.height * scale)
            img.width = int(img.width * scale)
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 15, 45)
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 3
    except Exception as e:
        logging.warning(f"Logó betöltése nem sikerült: {e}")


def month_to_year_and_quarter(selected_month_norm: str) -> tuple[int, int]:
    now = datetime.now()
    m = selected_month_norm.strip().lower()
    month_num = HONAPOK.index(m) + 1 if m in HONAPOK else now.month
    year = now.year
    m_obj = re.search(r"(20\d{2})", COMPLIANCE_FILE)
    if m_obj:
        year = int(m_obj.group(1))
    quarter = HONAP_TO_Q.get(month_num, (month_num - 1) // 3 + 1)
    return year, quarter


def autosize_columns(ws, min_col: int, max_col: int, min_row: int, max_row: int):
    for c in range(min_col, max_col + 1):
        width = 0
        for r in range(min_row, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                width = max(width, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 90)

def load_active_clients() -> set[str]:
    df = pd.read_excel(COMPLIANCE_FILE, sheet_name=CEGADATOK_SHEET)
    active_clients = set(
        df[df["Ügyfél aktív"].astype(str).str.strip().str.lower() == "igen"]["Ügyfélkód"].astype(str)
    )
    return active_clients


def load_client_name_map() -> dict[str, str]:
    df = pd.read_excel(COMPLIANCE_FILE, sheet_name=CEGADATOK_SHEET)
    df = df[df["Ügyfél aktív"].astype(str).str.strip().str.lower() == "igen"]
    if "Ügyfélkód" not in df.columns:
        raise ValueError("A Cégadatok lapon nincs 'Ügyfélkód' oszlop.")
    name_aliases = ["Cégnév", "Cég neve", "Ügyfél neve", "Partner neve", "Név"]
    name_col = next((c for c in name_aliases if c in df.columns), None) or "Ügyfélkód"
    df = df[["Ügyfélkód", name_col]].dropna(subset=["Ügyfélkód"])
    df["Ügyfélkód"] = df["Ügyfélkód"].astype(str)
    df[name_col] = df[name_col].astype(str)
    return dict(zip(df["Ügyfélkód"], df[name_col]))


def load_client_lang_map() -> dict[str, str]:
    df = pd.read_excel(COMPLIANCE_FILE, sheet_name=CEGADATOK_SHEET)
    df = df[df["Ügyfél aktív"].astype(str).str.strip().str.lower() == "igen"]

    # ha nincs Nyelv oszlop, fallback = "magyar"
    if "Nyelv" not in df.columns:
        return {str(k): "magyar" for k in df["Ügyfélkód"].astype(str)}

    df["Ügyfélkód"] = df["Ügyfélkód"].astype(str)
    df["Nyelv"] = df["Nyelv"].astype(str).str.strip().str.lower()

    return dict(zip(df["Ügyfélkód"], df["Nyelv"]))

# ---- FŐ FÜGGVÉNY ----
def generate_szamlamelleklet(
    selected_month: str, ordered_codes: list[str] | None = None
) -> str:
    """
    selected_month: 'januar'...'december' (ékezet nélkül)
    ordered_codes:  ha megadod, CSAK ezekre készül lap, ilyen sorrendben.
                    ha None, az ORDERED_CODES_DEFAULT lesz az alap (és csak azok, amelyek léteznek a Cégadatokban).

    A táblázat mostantól a 'Munka leírása' oszlop alapján gyűjt és összegez.
    """
    selected_month = remove_accents(selected_month.lower())
    logging.info(f"Hónap: {selected_month}")

    # 1) ügyfélkód + cégnév-térkép
    client_name_map_all = load_client_name_map()

    active = set(client_name_map_all.keys())  # már szűrt Cégadatok alapján

    codes_ordered: list[str]

    if ordered_codes:
        # DEFAULT_CLIENT_CODES szűrése aktív ügyfelekre
        codes_ordered = [c for c in ordered_codes if c in active]
    else:
        # default listából is csak aktív
        codes_ordered = [c for c in ORDERED_CODES_DEFAULT if c in active]

    client_name_map = {c: client_name_map_all[c] for c in codes_ordered}

    lang_map_all = load_client_lang_map()
    client_lang_map = {c: lang_map_all[c] for c in codes_ordered}

    # 2) adatgyűjtés: CSAK a kiválasztott kódokra
    #    ➜ kulcs: ügyfélkód -> { leírás -> össz_óra }
    description_summary: dict[str, dict[str, float]] = {
        kod: {} for kod in codes_ordered
    }

    DESCRIPTION_ALIASES = [
        "Munka leírása",
        "Feladat leírása",
        "Leírás",
        "Munka leirasa",
        "Leiras",
    ]
    HOURS_COL = "Időráfordítás (óra)"
    CLIENT_COL = "Ügyfélkód"
    # If a timesheet summary workbook exists, prefer using the latest one
    try:
        p = Path(FOLDER_PATH)
    except Exception:
        p = Path(".")

    summary_candidates = sorted(
        list(p.glob("timesheet_summary_*.xlsx")), key=lambda x: x.stat().st_mtime, reverse=True
    )
    used_summary = None
    if summary_candidates:
        for cand in summary_candidates:
            try:
                logging.info(f"Próbálom betölteni a timesheet összesítést: {cand}")
                # try expected sheet name first
                try:
                    df_sum = pd.read_excel(cand, sheet_name="Összesítés")
                except Exception:
                    # fallback to first sheet
                    df_sum = pd.read_excel(cand, sheet_name=0)

                # normalize column names to check presence
                cols = list(df_sum.columns)
                norm_cols = {remove_accents(c).strip().lower(): c for c in cols}
                needed_norm = [remove_accents(CLIENT_COL).lower(), remove_accents("Munka leírása").lower(), remove_accents(HOURS_COL).lower()]
                if all(n in norm_cols for n in needed_norm):
                    # map back to original columns
                    client_col_real = norm_cols[remove_accents(CLIENT_COL).lower()]
                    desc_col_real = norm_cols[remove_accents("Munka leírása").lower()]
                    hours_col_real = norm_cols[remove_accents(HOURS_COL).lower()]

                    # filter and aggregate
                    df_sum = df_sum[[client_col_real, desc_col_real, hours_col_real]]
                    df_sum.dropna(subset=[client_col_real, desc_col_real, hours_col_real], inplace=True)
                    df_sum[client_col_real] = df_sum[client_col_real].astype(str)
                    df_sum[desc_col_real] = df_sum[desc_col_real].astype(str)
                    # group and fill description_summary
                    grp = (
                        df_sum.groupby([client_col_real, desc_col_real], dropna=False)[hours_col_real]
                        .sum()
                        .reset_index()
                    )
                    for _, r in grp.iterrows():
                        kod = str(r[client_col_real])
                        if kod not in description_summary:
                            continue
                        desc = str(r[desc_col_real]).strip()
                        try:
                            hrs = float(r[hours_col_real])
                        except Exception:
                            continue
                        description_summary[kod][desc] = description_summary[kod].get(desc, 0.0) + hrs

                    used_summary = cand
                    logging.info(f"Használva timesheet összesítés: {cand}")
                    break
                else:
                    logging.warning(f"Nem található a szükséges oszlop a summary fájlban: {cand}")
            except Exception as e:
                logging.exception(f"Hiba timesheet összesítés beolvasásakor ({cand}): {e}")

    # If we didn't find/consume a valid summary, fall back to scanning individual TS files
    if used_summary is None:
        for file in os.listdir(FOLDER_PATH):
            if file.endswith(".xlsx") and "TS" in file and not file.startswith("~$"):
                path = os.path.join(FOLDER_PATH, file)
                logging.info(f"Feldolgozás: {file}")
                try:
                    xls = pd.ExcelFile(path)
                except Exception as e:
                    logging.exception(f"Nem nyitható: {file} — {e}")
                    continue

                for sheet in xls.sheet_names:
                    if remove_accents(str(sheet).lower()) != selected_month:
                        continue
                    try:
                        # Először megpróbáljuk közvetlenül a várt oszlopokkal
                        try:
                            df = pd.read_excel(
                                xls,
                                sheet_name=sheet,
                                usecols=[CLIENT_COL, DESCRIPTION_ALIASES[0], HOURS_COL],
                                nrows=MAX_ROWS_PER_SHEET,
                            )
                            desc_col = DESCRIPTION_ALIASES[0]
                        except Exception:
                            # Ha a pontos "Munka leírása" nincs, beolvassuk szélesen és kiválasztunk egy alias-t
                            df = pd.read_excel(
                                xls, sheet_name=sheet, nrows=MAX_ROWS_PER_SHEET
                            )
                            # normalizált név-térkép
                            cols_norm = {
                                c: remove_accents(str(c)).lower().strip()
                                for c in df.columns
                            }
                            # órák és ügyfélkód jelenléte
                            if CLIENT_COL not in df.columns:
                                raise KeyError(
                                    f"Hiányzik a '{CLIENT_COL}' oszlop: {file}/{sheet}"
                                )
                            if HOURS_COL not in df.columns:
                                raise KeyError(
                                    f"Hiányzik a '{HOURS_COL}' oszlop: {file}/{sheet}"
                                )
                            # keresünk leírás alias-t
                            desc_col = None
                            for cand2 in DESCRIPTION_ALIASES:
                                if cand2 in df.columns:
                                    desc_col = cand2
                                    break
                                # fallback: akcentus nélküli egyezés
                                cand_norm = remove_accents(cand2).lower()
                                for orig, norm in cols_norm.items():
                                    if norm == remove_accents(cand2).lower():
                                        desc_col = orig
                                        break
                                if desc_col:
                                    break
                            if not desc_col:
                                raise KeyError(
                                    f"Nem található 'Munka leírása' (alias-ok: {', '.join(DESCRIPTION_ALIASES)}) a {file}/{sheet} lapon."
                                )
                            df = df[[CLIENT_COL, desc_col, HOURS_COL]]
                    except Exception as e:
                        logging.exception(f"Hiba a sheet olvasásakor ({file}/{sheet}): {e}")
                        continue

                    df.dropna(subset=[CLIENT_COL, desc_col, HOURS_COL], inplace=True)

                    # típusok
                    df[CLIENT_COL] = df[CLIENT_COL].astype(str)
                    df[desc_col] = df[desc_col].astype(str)

                    for _, r in df.iterrows():
                        kod = str(r[CLIENT_COL])
                        if kod not in description_summary:  # nem kiválasztott ügyfél
                            continue
                        desc = str(r[desc_col]).strip()
                        try:
                            hrs = float(r[HOURS_COL])
                        except Exception:
                            continue
                        description_summary[kod][desc] = (
                            description_summary[kod].get(desc, 0.0) + hrs
                        )

    # 3) kimeneti excel — egyoszlopos layout, logó bal felül
    # ensure we don't overwrite an existing file for the same month
    base_out = Path(f"szamlamelleklet_{selected_month}.xlsx")
    if base_out.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"szamlamelleklet_{selected_month}_{ts}.xlsx"
        logging.info(f"Output file exists, writing to new file: {out_name}")
    else:
        out_name = str(base_out)
    wb = Workbook()
    used_titles = set()
    if wb.active is not None:
        wb.remove(wb.active)

    year = datetime.now().year
    month_num = (
        HONAPOK.index(selected_month) + 1
        if selected_month in HONAPOK
        else datetime.now().month
    )
    today_str = datetime.now().strftime("%Y. %m. %d.")

    # create sheets in alphabetical order (accent-insensitive) regardless of input order
    for kod in sorted(codes_ordered, key=lambda k: remove_accents(k).lower()):
        items = description_summary.get(kod, {})
        ws = wb.create_sheet(title=sanitize_sheet_title(kod, used_titles))

        # Logó bal felül + padding
        place_logo_top_left(ws)

        # Fejléc egy oszlopban (C)
        ws["C6"].value = client_name_map.get(kod, kod)
        ws["C6"].font = Font(italic=True, color="C00000", size=12)
        ws["C6"].alignment = Alignment(horizontal="left", vertical="center")
        ws["C6"].border = BORDER_BOTTOM_THICK

        lang = client_lang_map.get(kod, "magyar").lower()
        bp = BOILERPLATE.get(lang, BOILERPLATE["magyar"])

        ws["C10"].value = bp["title"]
        ws["C10"].font = Font(bold=True, size=14)
        ws["C11"].value = bp["period"].format(year=year, month=month_num)
        ws["C12"].value = bp["date"].format(today=today_str)
        ws["C12"].font = Font(italic=True, color="C00000")
        ws["C13"].value = bp["text"]
        for a in ("C10", "C11", "C12", "C13"):
            ws[a].alignment = Alignment(horizontal="left", wrap_text=True)

        # Tábla (C–D) — mostantól a 'Munka leírása' értékekkel
        ws["C16"].value = bp["task_header"]
        ws["D16"].value = bp["hours_header"]
        for cell in (ws["C16"], ws["D16"]):
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center")

        row = 17
        total = 0.0

        for desc, hrs in sorted(items.items(), key=lambda x: str(x[0]).lower()):
            ws.cell(row=row, column=3, value=str(desc)).alignment = Alignment(
                wrap_text=True
            )
            d = ws.cell(row=row, column=4, value=round(float(hrs), 2))
            d.number_format = "0.00"
            d.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=3).border = BORDER_THIN
            ws.cell(row=row, column=4).border = BORDER_THIN
            total += float(hrs)
            row += 1

        
        # Összegző sorok
        labels = bp["summary_labels"]
        values = [round(total, 2), 0, 0, None]
        first_task_row = 17
        last_task_row = row - 1


        # Összeg képlet
        sum_formula = f"=SUM(D{first_task_row}:D{last_task_row})"

        # Sor 1: Felhasznált órák
        cell = ws.cell(row=row, column=3, value=labels[0])
        cell.border = BORDER_THIN
        cell.font = Font(bold=True)
        cell1 = ws.cell(row=row, column=4, value=sum_formula)
        cell1.number_format = "0.00"
        cell1.border = BORDER_THIN
        

        # Sor 2
        cell = ws.cell(row=row+1, column=3, value=labels[1])
        cell.border = BORDER_THIN
        cell.font = Font(bold=True)
        cell1 = ws.cell(row=row+1, column=4, value=0)
        cell1.number_format = "0.00"
        cell1.border = BORDER_THIN

        # Sor 3
        cell = ws.cell(row=row+2, column=3, value=labels[2])
        cell.border = BORDER_THIN
        cell.font = Font(bold=True)
        cell1 = ws.cell(row=row+2, column=4, value=0)
        cell1.number_format = "0.00"
        cell1.border = BORDER_THIN

        # Sor 4: Különbözet
        cell = ws.cell(row=row+3, column=3, value=labels[3])
        cell.border = BORDER_THIN
        cell.font = Font(bold=True)
        cell1 = ws.cell(row=row+3, column=4, value=f"=D{row}-D{row+1}+D{row+2}")
        cell1.number_format = "0.00"
        cell1.border = BORDER_THIN


        # szélességek + auto magasság
        ws.column_dimensions["C"].width = 70
        ws.column_dimensions["D"].width = 18
        last_row = row + 3
        for r in range(6, last_row + 1):
            ws.row_dimensions[r].height = None
        autosize_columns(ws, min_col=1, max_col=7, min_row=6, max_row=last_row)

    wb.save(out_name)
    logging.info(f"Kész: {out_name}")
    return out_name


if __name__ == "__main__":
    # Konzolos használatnál: a második és további argumentumok lehetnek ügyfélkódok
    # pl.: python generate_szamlamelleklet.py januar AUC AXM MES
    if len(sys.argv) > 1:
        month = sys.argv[1]
        codes = sys.argv[2:] if len(sys.argv) > 2 else None
    else:
        month = datetime.now().strftime("%B").lower()
        codes = None
    month = remove_accents(month)
    try:
        print("Kész:", generate_szamlamelleklet(month, ordered_codes=codes))
    except Exception as e:
        print("Hiba:", e)
        sys.exit(1)
