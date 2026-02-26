import unicodedata
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def norm_header(s):
    """Normalizes strings for comparison (removes accents, lowercase, strips)."""
    if not isinstance(s, str):
        return str(s)
    return (
        "".join(
            c
            for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        )
        .lower()
        .strip()
    )


def autosize_columns(sheet):
    """Automatically adjusts column widths based on content."""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column].width = max_length + 2


def add_title_banner(sheet, title, month, col_count=7):
    """Adds the standard Ecovis styled title banner to a sheet."""
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = sheet.cell(row=1, column=1)
    cell.value = f"{title} - {month}"
    cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    sheet.row_dimensions[1].height = 30


def write_table(sheet, df, start_row, start_col=1):
    """Writes a DataFrame to a sheet with standard styling and borders."""
    # Header styling
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for j, col_name in enumerate(df.columns):
        cell = sheet.cell(row=start_row, column=start_col + j)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            cell = sheet.cell(row=start_row + i + 1, column=start_col + j)
            cell.value = value
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
